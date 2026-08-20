# -*- coding: utf-8 -*-
"""
ETL segunda pasada: beneficio seco, existencias, muestras y ventas.

Lee los archivos de SALDO DE ALMACENES y TRAZABILIDAD DE VENTA, que traen el
desglose de cada lote por productor. A diferencia del acopio, aqui los
productores aparecen SOLO por nombre escrito a mano, sin codigo: el
emparejamiento se hace por nombre normalizado contra db/out/personas.csv.

    python scripts/etl_beneficio.py

Requiere haber corrido antes scripts/etl_excel.py (usa sus CSV para resolver
personas y lotes).
"""

import argparse
import csv
import os
import re
import sys
import unicodedata
import uuid
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl.  pip install openpyxl")

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# (fuente, archivo, hoja). La fuente define la semantica de las filas.
FUENTES = [
    ("saldo_almacen", "SALDO DE ALMACENES LA PAZ/SALDO DE CAFE EN TRANSICION 2024.xlsx", "SALDOS TRANS_25"),
    ("saldo_almacen", "SALDO DE ALMACENES LA PAZ/SALDO DE CAFE ORGANICO 2024.xlsx", "SALDO DE CAFE ORG_25"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE EN TRANSICION.xlsx", "EX_KOREA"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE EN TRANSICION.xlsx", "Hoja1"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE ORGANICO.xlsx", "VENTA PLANTA EL ALTO 1"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE ORGANICO.xlsx", "hoja_1"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE ORGANICO.xlsx", "hoja_2"),
    ("trazabilidad_venta", "TRAZAVILIDAD DE VENTA DE CAFE 2024/TRAZABILIDAD DE CAFE ORGANICO.xlsx", "hoja_3"),
]

IGNORAR = {"SUB TOTAL", "TOTAL", "TOTALES", "NOMBRES Y APELLIDOS", ""}

# Ventas identificadas a mano contra la celda de origen. No se parsean porque
# cada hoja tiene un layout distinto y adivinar inventaria trazabilidad.
# Ver docs: solo se incluyen las que se pudieron anclar a lotes concretos.
VENTAS = [
    {
        "numero": "ANDES-2025-09",
        "cliente": "ANDES COFFEE",
        "fecha": "2025-09-19",
        "total_kg": 14100,
        "almacen": "Almacen El Alto",
        "origen": "EX_KOREA, bloque 'VENTA DE LOTE ANDES COFFEE'",
        # Cada lote con sus kg explicitos en la hoja.
        "lotes": {"TR-02-24": 1683.0688, "OR-07-24": 699.2, "OR-01-25": 4410.72,
                  "OR-02-25": 4094.72, "TR-01-25": 3214.62},
        "muestras": [("OR-01-25", "muestra", 2.3288)],
    },
    {
        "numero": "ANNK-2026-02",
        "cliente": "Ann Katterine",
        "fecha": "2026-02-26",
        "sacos": 300,
        "kg_por_saco": 60,
        "total_kg": 18000,
        "almacen": "Almacen El Alto",
        "origen": "VENTA PLANTA EL ALTO 1, bloque CONTRATO 1",
        # LOTE 3 y LOTE 4 confirmados por aritmetica exacta contra el verde
        # del lote: 30.390 * 0,16 = 4.862,4 y 28.680 * 0,16 = 4.588,8.
        "lotes": {"OR-03-25": 4862.4, "OR-04-25": 4588.8, "OR-05-25": 4239.0},
        "muestras": [("OR-05-25", "muestra", 5.0), ("OR-05-25", "contramuestra", 5.0)],
    },
]


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().upper()


def norm_lote(c):
    """OR-3-25, LOTE TR-01, TR-02-24 -> OR-03-25 / TR-01-25 / TR-02-24."""
    if not c:
        return None
    c = re.sub(r"\s+", "", str(c)).upper().replace("LOTE", "")
    m = re.match(r"^(OR|TR)-?(\d+)(?:-(\d+))?$", c)
    return f"{m.group(1)}-{int(m.group(2)):02d}-{m.group(3) or '25'}" if m else None


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def leer_csv(nombre):
    ruta = os.path.join(RAIZ, "db", "out", nombre + ".csv")
    if not os.path.exists(ruta):
        sys.exit(f"Falta {ruta}. Corre primero: python scripts/etl_excel.py")
    with open(ruta, encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def escribir(out, nombre, filas, columnas):
    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, nombre + ".csv"), "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columnas, extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)
    print(f"  {nombre + '.csv':28} {len(filas):5} filas")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default=os.path.join(RAIZ, "trazabilidad 2025"))
    ap.add_argument("--out", default=os.path.join(RAIZ, "db", "out"))
    args = ap.parse_args()

    # --- contexto de la primera pasada
    personas = {norm(p["nombre"]): p["id"] for p in leer_csv("personas")}
    lotes = {l["codigo"]: dict(l) for l in leer_csv("lotes")}
    acopio_por_lote = defaultdict(float)
    id_a_codigo = {l["id"]: l["codigo"] for l in lotes.values()}
    for e in leer_csv("entregas_acopio"):
        cod = id_a_codigo.get(e["lote_id"])
        if cod:
            acopio_por_lote[cod] += float(e["kg_guinda_real"])

    # --- leer el desglose por productor
    print("Leyendo archivos de saldo y venta...")
    filas = []
    for fuente, rel, hoja in FUENTES:
        ruta = os.path.join(args.base, rel)
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        if hoja in wb.sheetnames:
            for r in wb[hoja].iter_rows(min_row=5, values_only=True):
                if not r or len(r) < 10:
                    continue
                nombre = str(r[1]).strip() if r[1] else ""
                if norm(nombre) in IGNORAR:
                    continue
                cod = norm_lote(r[9])
                if not cod:
                    continue
                filas.append({
                    "fuente": fuente, "hoja": hoja, "nombre_excel": nombre, "lote": cod,
                    "latas": num(r[2]), "kg_guinda": num(r[3]), "kg_pergamino_seco": num(r[4]),
                    "kg_trillado": num(r[5]), "kg_descarte": num(r[6]),
                    "kg_caracol": num(r[7]), "kg_verde_export": num(r[8]),
                })
        wb.close()
    print(f"  {len(filas)} filas de desglose por productor")

    # --- lotes nuevos (gestiones anteriores referenciadas en las ventas)
    referidos = {f["lote"] for f in filas} | {c for v in VENTAS for c in v["lotes"]}
    nuevos = []
    for cod in sorted(referidos - set(lotes)):
        pre, numero, anio = cod.split("-")
        reg = {"id": str(uuid.uuid4()), "codigo": cod, "campania_id": 2000 + int(anio),
               "certificacion": "organico" if pre == "OR" else "transicion",
               "correlativo": int(numero)}
        lotes[cod] = reg
        nuevos.append(reg)
    if nuevos:
        print(f"  lotes de gestiones anteriores a crear: {[n['codigo'] for n in nuevos]}")

    # --- beneficio_seco: uno por lote, con las columnas _calc del Excel
    beneficio = {}
    agreg = defaultdict(lambda: defaultdict(float))
    for f in filas:
        for k in ("kg_pergamino_seco", "kg_trillado", "kg_verde_export", "kg_caracol", "kg_descarte"):
            agreg[f["lote"]][k] += f[k]
    for cod, a in sorted(agreg.items()):
        if a["kg_pergamino_seco"] <= 0:
            continue
        beneficio[cod] = {
            "id": str(uuid.uuid4()), "lote_id": lotes[cod]["id"],
            "kg_pergamino_entrada": round(a["kg_pergamino_seco"], 3),
            "kg_trillado_calc": round(a["kg_trillado"], 3),
            "kg_verde_calc": round(a["kg_verde_export"], 3),
            "kg_caracol_calc": round(a["kg_caracol"], 3),
            "kg_descarte_calc": round(a["kg_descarte"], 3),
            "observaciones": "Derivado del desglose por productor. Sin peso fisico medido.",
        }

    # --- beneficio_productor
    # Un lote "no cuadra" cuando lo que el desglose reparte no coincide con lo
    # que el acopio dice que entro. Se marca, no se corrige.
    detalle, sin_persona, no_cuadra = [], 0, set()
    kg_ben = defaultdict(float)
    for f in filas:
        kg_ben[f["lote"]] += f["kg_guinda"]
    for cod in kg_ben:
        base = acopio_por_lote.get(cod)
        if base and abs(kg_ben[cod] - base) > 0.5:
            no_cuadra.add(cod)

    for f in filas:
        pid = personas.get(norm(f["nombre_excel"]))
        notas, revision = [], "ok"
        if not pid:
            revision = "observado"
            notas.append(f"Nombre '{f['nombre_excel']}' sin coincidencia en el padron")
            sin_persona += 1
        if f["lote"] in no_cuadra:
            revision = "observado"
            notas.append(f"El desglose del lote {f['lote']} no reconcilia con el acopio "
                         f"({kg_ben[f['lote']]:.1f} vs {acopio_por_lote[f['lote']]:.1f} kg)")
        b = beneficio.get(f["lote"])
        if not b:
            continue
        detalle.append({
            "beneficio_id": b["id"], "persona_id": pid or "", "parcela_id": "",
            "nombre_excel": f["nombre_excel"], "fuente": f["fuente"],
            "latas": round(f["latas"], 4), "kg_guinda": round(f["kg_guinda"], 3),
            "kg_pergamino_seco": round(f["kg_pergamino_seco"], 3),
            "kg_trillado": round(f["kg_trillado"], 3),
            "kg_descarte": round(f["kg_descarte"], 3),
            "kg_caracol": round(f["kg_caracol"], 3),
            "kg_verde_export": round(f["kg_verde_export"], 3),
            "revision": revision, "revision_nota": " | ".join(notas),
        })

    # --- existencias: lo que las hojas de SALDO dicen que queda en bodega
    existencias = []
    stock = defaultdict(float)
    for f in filas:
        if f["fuente"] == "saldo_almacen":
            stock[f["lote"]] += f["kg_verde_export"]
    for cod, kg in sorted(stock.items()):
        if kg > 0:
            existencias.append({
                "lote_id": lotes[cod]["id"], "producto": "verde_oro",
                "kg_ingreso": round(kg, 3), "kg_saldo": round(kg, 3),
                "fecha_ingreso": "2025-12-31",
                "responsable": "Carga inicial desde SALDO DE ALMACENES",
            })

    # --- ventas
    contratos, despachos, desp_lotes, muestras, exportaciones = [], [], [], [], []
    for v in VENTAS:
        cid, did = str(uuid.uuid4()), str(uuid.uuid4())
        contratos.append({
            "id": cid, "numero": v["numero"], "cliente": v["cliente"], "fecha": v["fecha"],
            "sacos": v.get("sacos", ""), "kg_por_saco": v.get("kg_por_saco", ""),
            "puerto_destino": "", "moneda": "USD",
        })
        despachos.append({
            "id": did, "contrato_id": cid, "almacen": v["almacen"],
            "fecha_despacho": v["fecha"], "kg_neto": v["total_kg"],
            "observaciones": "Origen: " + v["origen"],
        })
        for cod, kg in v["lotes"].items():
            desp_lotes.append({"despacho_id": did, "lote_id": lotes[cod]["id"],
                               "kg_asignados": round(kg, 3)})
        for cod, tipo, kg in v.get("muestras", []):
            muestras.append({"lote_id": lotes[cod]["id"], "tipo": tipo, "kg": kg,
                             "fecha": v["fecha"], "motivo": "Control de calidad / cateo",
                             "responsable": ""})
        exportaciones.append({"despacho_id": did, "fecha_embarque": v["fecha"],
                              "volumen_kg": v["total_kg"]})

    # --- escribir
    print("\nEscribiendo CSV...")
    out = args.out
    escribir(out, "lotes_extra", nuevos, ["id", "codigo", "campania_id", "certificacion", "correlativo"])
    escribir(out, "beneficio_seco", list(beneficio.values()),
             ["id", "lote_id", "kg_pergamino_entrada", "kg_trillado_calc", "kg_verde_calc",
              "kg_caracol_calc", "kg_descarte_calc", "observaciones"])
    escribir(out, "beneficio_productor", detalle,
             ["beneficio_id", "persona_id", "parcela_id", "nombre_excel", "fuente", "latas",
              "kg_guinda", "kg_pergamino_seco", "kg_trillado", "kg_descarte", "kg_caracol",
              "kg_verde_export", "revision", "revision_nota"])
    escribir(out, "existencias", existencias,
             ["lote_id", "producto", "kg_ingreso", "kg_saldo", "fecha_ingreso", "responsable"])
    escribir(out, "contratos", contratos,
             ["id", "numero", "cliente", "fecha", "sacos", "kg_por_saco", "puerto_destino", "moneda"])
    escribir(out, "despachos", despachos,
             ["id", "contrato_id", "almacen", "fecha_despacho", "kg_neto", "observaciones"])
    escribir(out, "despacho_lotes", desp_lotes, ["despacho_id", "lote_id", "kg_asignados"])
    escribir(out, "muestras", muestras, ["lote_id", "tipo", "kg", "fecha", "motivo", "responsable"])
    escribir(out, "exportaciones", exportaciones, ["despacho_id", "fecha_embarque", "volumen_kg"])

    obs = sum(1 for d in detalle if d["revision"] != "ok")
    print(f"""
RESUMEN
  lotes nuevos (gestiones previas) . {len(nuevos)}
  beneficio_seco ................... {len(beneficio)} lotes
  beneficio_productor .............. {len(detalle)} filas
  existencias ...................... {len(existencias)}
  contratos / despachos ............ {len(contratos)}
  muestras ......................... {len(muestras)}

A REVISAR
  filas observadas ................. {obs} de {len(detalle)}
  nombres sin emparejar ............ {sin_persona}
  lotes que no reconcilian ......... {len(no_cuadra)} -> {sorted(no_cuadra)}""")


if __name__ == "__main__":
    main()
