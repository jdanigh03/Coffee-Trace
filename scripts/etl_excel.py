# -*- coding: utf-8 -*-
"""
ETL: Excel de trazabilidad -> CSV normalizados listos para Supabase.

No inserta en la base. Genera CSV en db/out/ para poder revisarlos antes de
cargar, que es lo sensato dado que los Excel traen inconsistencias conocidas.

Uso:
    python scripts/etl_excel.py
    python scripts/etl_excel.py --base "ruta/a/trazabilidad 2025" --out db/out

Cubre la cadena principal: personas, parcelas, codigos_productor,
certificaciones, lotes, entregas_acopio y envios.
El beneficio seco y las ventas quedan fuera a proposito: ahi los productores se
identifican solo por nombre escrito a mano, y emparejarlos automaticamente
mezclaria lotes de distintos socios. Ver README.
"""

import argparse
import csv
import os
import re
import sys
import unicodedata
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl.  pip install openpyxl")

KG_POR_LATA = 14

ARCHIVOS = {
    "acopio_org": ("ACOPIO DE GUINDA 2025/PLANILLA DE ACOPIO ORG.2025.xlsx", "ACOPIO ORG_2025"),
    "acopio_tra": ("ACOPIO DE GUINDA 2025/PLANILLA DE ACOPIO TRANS_25.xlsx", "TRANS_2025"),
    "seg_org":    ("seguimiento a cosecha 2025/SEGUIMIENTO DE LOTES DE CAFE organico2025.xlsx", None),
    "seg_tra":    ("seguimiento a cosecha 2025/SEGUIMINETO DE LOTES DE CAFE TRASICION2025.xlsx", None),
    "envio_org":  ("REGUISTRO DE ENVIO TAIPIPLAYA LA PAZ/REGISTRO DE ENVIO TAI_LA PAZ ORGANICO_25.xlsx",
                   "PLAN_ENVIO PERGAMINO"),
    "envio_tra":  ("REGUISTRO DE ENVIO TAIPIPLAYA LA PAZ/REGISTRO DE ENVIO DE CAFE TRANSICION_25.xlsx",
                   "ENVIO DE CAFE TRANS_24"),
}

# Los factores viven en la base (tabla factores_conversion). Aqui solo se usa
# el de pergamino para derivar el peso despachado del envio.
FACTOR_GUINDA_A_PERGAMINO = 0.20


def norm_nombre(s):
    """Normaliza para deduplicar: sin acentos, sin dobles espacios, mayusculas."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().upper()


def norm_codigo(s):
    return re.sub(r"\s+", "", str(s)).upper() if s else ""


def norm_lote(codigo, anio_default=25):
    """OR-01, OR-3-25, TR-02-24 -> OR-01-25 / TR-02-24."""
    if not codigo:
        return None
    c = re.sub(r"\s+", "", str(codigo)).upper().replace("LOTE", "")
    m = re.match(r"^(OR|TR)-?(\d+)(?:-(\d+))?$", c)
    if not m:
        return None
    pre, num, anio = m.group(1), int(m.group(2)), m.group(3)
    return f"{pre}-{num:02d}-{anio or anio_default}"


def leer(base, rel, hoja):
    wb = openpyxl.load_workbook(os.path.join(base, rel), data_only=True, read_only=True)
    filas = list(wb[hoja].iter_rows(values_only=True))
    wb.close()
    return filas


def hojas(base, rel):
    wb = openpyxl.load_workbook(os.path.join(base, rel), data_only=True, read_only=True)
    nombres = wb.sheetnames
    wb.close()
    return nombres


def a_fecha(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


# ------------------------------------------------------------------
# Lectura
# ------------------------------------------------------------------

def leer_acopio(base):
    """Devuelve las entregas crudas de ambas planillas."""
    filas = []
    for clave, cert in (("acopio_org", "organico"), ("acopio_tra", "transicion")):
        rel, hoja = ARCHIVOS[clave]
        for r in leer(base, rel, hoja)[1:]:
            if not r or r[0] is None or r[1] is None:
                continue
            fecha = a_fecha(r[0])
            if not fecha:
                continue
            filas.append({
                "fecha": fecha,
                "codigo": norm_codigo(r[1]),
                "nombre_excel": str(r[2]).strip() if r[2] else "",
                "comunidad": str(r[3]).strip() if r[3] else "",
                "kg": float(r[4]) if r[4] else 0.0,
                "precio": float(r[6]) if r[6] else 0.0,
                "estatus": str(r[8]).strip().upper() if r[8] else None,
                "archivo": cert,
            })
    return filas


def leer_asignacion_lotes(base):
    """
    Lee las hojas OR-NN / TR-NN y devuelve dos mapas:

      exacto : (fecha, codigo, kg) -> lote
      laxo   : (fecha, codigo)     -> (lote, kg_segun_seguimiento)

    Hacen falta los dos porque desde el 2025-07-20 la planilla de acopio y el
    archivo de seguimiento discrepan en los kilos de cada entrega: coinciden la
    fecha y el codigo, pero no el peso. El mapa laxo permite asignar el lote
    igual y dejar la discrepancia marcada en vez de perder la entrega.
    """
    exacto, laxo = {}, {}
    for clave, pre in (("seg_org", "OR"), ("seg_tra", "TR")):
        rel, _ = ARCHIVOS[clave]
        for hoja in hojas(base, rel):
            if not re.match(rf"^{pre}-\d+$", hoja):
                continue
            lote = norm_lote(hoja)
            for r in leer(base, rel, hoja)[1:]:
                if not r or r[0] is None or r[1] is None:
                    continue
                fecha = a_fecha(r[0])
                if not fecha or not r[4]:
                    continue
                cod, kg = norm_codigo(r[1]), round(float(r[4]), 2)
                exacto[(fecha, cod, kg)] = lote
                laxo.setdefault((fecha, cod), (lote, kg))
    return exacto, laxo


def leer_envios(base):
    envios = []
    for clave, col0 in (("envio_org", 2), ("envio_tra", 1)):
        rel, hoja = ARCHIVOS[clave]
        for r in leer(base, rel, hoja)[5:]:
            if not r or len(r) <= col0 + 12:
                continue
            fecha = a_fecha(r[col0])
            guinda = r[col0 + 2]
            lote = r[col0 + 12]
            if not fecha or not guinda or not lote:
                continue
            envios.append({
                "lote": norm_lote(lote),
                "fecha_salida": fecha,
                # El Excel registra guinda; lo que viaja es pergamino.
                "kg_pergamino_despachado": round(float(guinda) * FACTOR_GUINDA_A_PERGAMINO, 3),
            })
    return envios


# ------------------------------------------------------------------
# Normalizacion
# ------------------------------------------------------------------

def construir(filas, mapas_lote, envios, campania):
    mapa_exacto, mapa_laxo = mapas_lote
    # -- personas: una por nombre normalizado
    personas = {}
    for f in filas:
        k = norm_nombre(f["nombre_excel"])
        if k and k not in personas:
            personas[k] = {"id": str(uuid.uuid4()), "nombre": f["nombre_excel"].strip()}

    # -- parcelas: una por codigo. El titular es el nombre mas frecuente de ese codigo.
    nombres_por_codigo = defaultdict(Counter)
    comunidad_por_codigo = defaultdict(Counter)
    estatus_por_codigo = defaultdict(Counter)
    for f in filas:
        if f["nombre_excel"]:
            nombres_por_codigo[f["codigo"]][norm_nombre(f["nombre_excel"])] += 1
        if f["comunidad"]:
            comunidad_por_codigo[f["codigo"]][f["comunidad"]] += 1
        if f["estatus"]:
            estatus_por_codigo[f["codigo"]][f["estatus"]] += 1

    parcelas, codigos = {}, {}
    for cod, cnt in nombres_por_codigo.items():
        titular = cnt.most_common(1)[0][0]
        pid = str(uuid.uuid4())
        parcelas[cod] = {
            "id": pid,
            "persona_id": personas[titular]["id"],
            "comunidad": comunidad_por_codigo[cod].most_common(1)[0][0] if comunidad_por_codigo[cod] else "",
            "nombres_distintos": len(cnt),
        }
        codigos[cod] = {"id": str(uuid.uuid4()), "codigo": cod, "parcela_id": pid,
                        "persona_id": personas[titular]["id"],
                        "nombre_excel": cnt.most_common(1)[0][0]}

    # -- certificaciones: el estatus es constante por codigo (verificado)
    certificaciones = []
    for cod, cnt in estatus_por_codigo.items():
        est = cnt.most_common(1)[0][0]
        certificaciones.append({
            "parcela_id": parcelas[cod]["id"],
            "campania_id": campania,
            "estatus": est,
            "tipo": "organico" if est == "E" else "transicion",
            "estatus_distintos": len(cnt),
        })

    # -- lotes
    codigos_lote = sorted({v for v in mapa_exacto.values() if v} |
                          {e["lote"] for e in envios if e["lote"]})
    lotes = {}
    for cl in codigos_lote:
        pre, num, anio = cl.split("-")
        lotes[cl] = {
            "id": str(uuid.uuid4()),
            "codigo": cl,
            "campania_id": 2000 + int(anio),
            "certificacion": "organico" if pre == "OR" else "transicion",
            "correlativo": int(num),
        }

    # -- entregas
    entregas, obs = [], Counter()
    for f in filas:
        cod = f["codigo"]
        kg = round(f["kg"], 2)
        revision, notas = "ok", []

        lote_cod = mapa_exacto.get((f["fecha"], cod, kg))
        if lote_cod is None:
            # Sin coincidencia exacta: se asigna por fecha+codigo y se deja
            # constancia de que los dos archivos no dicen el mismo peso.
            laxo = mapa_laxo.get((f["fecha"], cod))
            if laxo:
                lote_cod, kg_seg = laxo
                revision = "observado"
                notas.append(f"Peso discrepante entre archivos: planilla de acopio {kg} kg vs "
                             f"seguimiento {kg_seg} kg. Se conserva el de acopio, que es el "
                             f"que se uso para pagar.")
                obs["peso_discrepante"] += 1
        lote = lotes.get(lote_cod) if lote_cod else None

        if not f["estatus"]:
            revision = "observado"
            notas.append("Estatus vacio en la planilla")
            obs["estatus_vacio"] += 1

        if f["estatus"]:
            tipo_prod = "organico" if f["estatus"] == "E" else "transicion"
            if tipo_prod != f["archivo"]:
                revision = "observado"
                notas.append(f"Estatus {f['estatus']} dentro de la planilla {f['archivo']}. "
                             "Cafe de transicion no puede pasar a organico.")
                obs["estatus_vs_planilla"] += 1
            if lote and tipo_prod != lote["certificacion"]:
                revision = "observado"
                notas.append(f"Estatus {f['estatus']} pero el lote {lote['codigo']} es "
                             f"{lote['certificacion']}")
                obs["estatus_vs_lote"] += 1

        if not lote:
            obs["sin_lote"] += 1

        entregas.append({
            "campania_id": campania,
            "fecha": f["fecha"],
            "codigo_productor_id": codigos[cod]["id"],
            "parcela_id": parcelas[cod]["id"],
            "persona_id": codigos[cod]["persona_id"],
            "kg_guinda_real": round(f["kg"], 2),
            "precio_unitario_bs": round(f["precio"], 2),
            "estatus_declarado": f["estatus"] or "",
            "lote_id": lote["id"] if lote else "",
            "revision": revision,
            "revision_nota": " | ".join(notas),
        })

    envios_out = []
    for e in envios:
        l = lotes.get(e["lote"])
        if l:
            envios_out.append({
                "lote_id": l["id"],
                "fecha_salida": e["fecha_salida"],
                "kg_pergamino_despachado": e["kg_pergamino_despachado"],
            })

    return personas, parcelas, codigos, certificaciones, lotes, entregas, envios_out, obs


# ------------------------------------------------------------------

def escribir(out, nombre, filas, columnas):
    os.makedirs(out, exist_ok=True)
    ruta = os.path.join(out, nombre + ".csv")
    with open(ruta, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=columnas, extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)
    print(f"  {nombre + '.csv':28} {len(filas):5} filas")
    return ruta


def main():
    aqui = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default=os.path.join(aqui, "trazabilidad 2025"))
    ap.add_argument("--out", default=os.path.join(aqui, "db", "out"))
    ap.add_argument("--campania", type=int, default=2025)
    args = ap.parse_args()

    if not os.path.isdir(args.base):
        sys.exit(f"No existe la carpeta: {args.base}")

    print("Leyendo Excel...")
    filas = leer_acopio(args.base)
    mapa_exacto, mapa_laxo = leer_asignacion_lotes(args.base)
    envios = leer_envios(args.base)
    print(f"  entregas={len(filas)}  asignaciones_lote={len(mapa_exacto)}  envios={len(envios)}")

    personas, parcelas, codigos, certs, lotes, entregas, envios_out, obs = construir(
        filas, (mapa_exacto, mapa_laxo), envios, args.campania)

    print("\nEscribiendo CSV...")
    escribir(args.out, "personas", list(personas.values()), ["id", "nombre"])
    escribir(args.out, "parcelas",
             [{"id": p["id"], "persona_id": p["persona_id"], "comunidad": p["comunidad"]}
              for p in parcelas.values()],
             ["id", "persona_id", "comunidad"])
    escribir(args.out, "codigos_productor", list(codigos.values()),
             ["id", "codigo", "parcela_id", "persona_id", "nombre_excel"])
    escribir(args.out, "certificaciones", certs,
             ["parcela_id", "campania_id", "estatus", "tipo"])
    escribir(args.out, "lotes", list(lotes.values()),
             ["id", "codigo", "campania_id", "certificacion", "correlativo"])
    escribir(args.out, "entregas_acopio", entregas,
             ["campania_id", "fecha", "codigo_productor_id", "parcela_id", "persona_id",
              "kg_guinda_real", "precio_unitario_bs", "estatus_declarado", "lote_id",
              "revision", "revision_nota"])
    escribir(args.out, "envios", envios_out,
             ["lote_id", "fecha_salida", "kg_pergamino_despachado"])

    # ---- control de calidad
    total_kg = sum(e["kg_guinda_real"] for e in entregas)
    total_pagado = sum(e["kg_guinda_real"] / KG_POR_LATA * e["precio_unitario_bs"] for e in entregas)
    multi_nombre = sum(1 for p in parcelas.values() if p["nombres_distintos"] > 1)
    multi_estatus = sum(1 for c in certs if c["estatus_distintos"] > 1)

    print(f"""
RESUMEN
  personas unicas .............. {len(personas)}
  parcelas (1 por codigo) ...... {len(parcelas)}
  lotes ........................ {len(lotes)}
  entregas ..................... {len(entregas)}
  kg de guinda ................. {total_kg:,.2f}
  total pagado (Bs) ............ {total_pagado:,.2f}

A REVISAR
  entregas sin lote asignado ... {obs['sin_lote']}
  peso discrepante entre archivos {obs['peso_discrepante']}
  estatus vacio ................ {obs['estatus_vacio']}
  estatus vs planilla .......... {obs['estatus_vs_planilla']}
  estatus vs lote .............. {obs['estatus_vs_lote']}
  codigos con >1 nombre ........ {multi_nombre}
  codigos con >1 estatus ....... {multi_estatus}

Los CSV estan en {args.out}
Revisalos antes de cargar: db/out/entregas_acopio.csv trae la columna `revision`.""")


if __name__ == "__main__":
    main()
